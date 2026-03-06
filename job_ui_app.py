import streamlit as st
import requests
import pandas as pd
import re
from access_lock import verify_access
from urllib.parse import quote
from datetime import datetime, timedelta
from info_panel import show_getting_started_panel

verify_access()

if "token" not in st.session_state:
    st.session_state["token"] = st.query_params.get("token")

st.set_page_config(
    page_title="Global Job Aggregator",
    layout="wide",
    initial_sidebar_state="expanded"
)

if "search_triggered" not in st.session_state:
    st.session_state["search_triggered"] = False

if "jobs_df" not in st.session_state:
    st.session_state["jobs_df"] = None

if "fallback_message" not in st.session_state:
    st.session_state["fallback_message"] = None

def rerun_search():
    if st.session_state.get("search_triggered"):
        st.session_state["rerun_needed"] = True

if "rerun_needed" not in st.session_state:
    st.session_state["rerun_needed"] = False

from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def export_to_excel(df):
    df_export = df.copy()

    # Remove Source column
    if "Source" in df_export.columns:
        df_export = df_export.drop(columns=["Source"])

    # Handle Apply columns
    if "Apply" in df_export.columns:
        df_export = df_export.rename(columns={"Apply": "Source Link"})
        df_export["Apply"] = df_export["Source Link"].apply(
            lambda x: f'=HYPERLINK("{x}","Apply")' if pd.notna(x) else ""
        )

    # Remove internal columns
    for col in ["API", "_excel", "_date"]:
        if col in df_export.columns:
            df_export = df_export.drop(columns=[col])

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="Jobs")
        ws = writer.sheets["Jobs"]

        # Header styling
        header_fill = PatternFill(start_color="4F6CF7", end_color="4F6CF7", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Freeze header row
        ws.freeze_panes = "A2"

        # Apply filter safely
        ws.auto_filter.ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"

        # Alternate row shading
        alt_fill = PatternFill(start_color="F4F6FF", end_color="F4F6FF", fill_type="solid")
        for row in range(2, ws.max_row + 1):
            if row % 2 == 0:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = alt_fill

        # Auto column sizing
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = max_length + 3

        # Make Source Link narrow
        for i, col_name in enumerate(df_export.columns, 1):
            if col_name == "Source Link":
                ws.column_dimensions[get_column_letter(i)].width = 8

    return output.getvalue()




# Keep your specific backend URL
BACKEND_URL ="https://global-job-aggregator-production.up.railway.app"


# ---------- RESTORED FRONT-END VISUALS ----------

# Manual sidebar button
st.sidebar.markdown("### ⭐ Help & Info")

with st.sidebar.expander("⭐ Getting Started & Refer", expanded=True):
    show_getting_started_panel()



st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@600;700;800&display=swap');

/* 1. TOP SPACE FIX */
.block-container {
    padding-top: 0rem !important;
    padding-bottom: 2rem !important;
}
header[data-testid="stHeader"] {
    background: transparent !important;
}

/* 2. THE CORRECT "START OF DAY" TITLE */
.hero-title {
    font-family: 'Inter', sans-serif !important;
    font-size: 52px !important;
    font-weight: 800 !important;
    line-height: 1.1;
    letter-spacing: -1px;
    text-align: center;
    margin: 0 auto;

    /* Restored the exact original 3-color gradient */
    background: linear-gradient(
        90deg, 
        #4F6CF7 0%, 
        #7A6FF0 50%, 
        #FF8A00 100%
    ) !important;

    -webkit-background-clip: text !important;
    -webkit-text-fill-color: transparent !important;
}

.hero-subtitle {
    font-family: 'Inter', sans-serif;
    font-size: 18px;
    color: #475569;
    margin-top: 10px;
    text-align: center;
}

/* 3. APP BACKGROUND & SIDEBAR */
.stApp {
    background: linear-gradient(135deg, #f5f3ff 0%, #fdf2f8 50%, #fff7ed 100%);
}
section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #6A5AE0, #B983FF) !important;
}
section[data-testid="stSidebar"] * {
    color: white !important;
}

/* 4. BUTTONS & CARDS */
.stButton>button, .apply-btn {
    background: linear-gradient(135deg, #FF5EDF, #FF8A00) !important;
    color: white !important;
    border-radius: 14px;
    font-weight: 600;
}
.job-card {
    background: rgba(255,255,255,0.92);
    border-radius: 18px;
    padding: 18px;
    box-shadow: 0 15px 35px rgba(0,0,0,0.08);
}

@media (max-width: 768px) {
    .hero-title { font-size: 32px !important; }
}
</style>

<div style="padding: 30px 0 20px 0; text-align:center;">
    <div class="hero-title">Global Job Aggregator</div>
    <div class="hero-subtitle">Search smarter. Apply faster.</div>
</div>
""", unsafe_allow_html=True)



st.markdown("""
<style>

/* EXPANDER HEADER STYLE */
[data-testid="stExpander"] details > summary {
    background: linear-gradient(135deg, #FF5EDF, #FF8A00) !important;
    color: white !important;
    border-radius: 10px;
    padding: 10px 14px;
    font-weight: 600;
}

[data-testid="stExpander"] details > summary div {
    background: transparent !important;
}

[data-testid="stExpander"] details > summary:hover {
    background: linear-gradient(135deg, #ff3fcf, #ff7200) !important;
}

[data-testid="stExpander"] svg {
    fill: white !important;
}

</style>
""", unsafe_allow_html=True)





st.markdown("""
<style>
/* Floating Help Button */
.floating-help-btn {
    position: fixed;
    bottom: 25px;
    right: 25px;
    z-index: 9999;
}

.floating-help-btn button {
    background: linear-gradient(135deg, #4F6CF7, #7A6FF0);
    color: white;
    border-radius: 50px;
    padding: 14px 22px;
    font-weight: 600;
    border: none;
    box-shadow: 0 10px 25px rgba(79,108,247,0.4);
    font-size: 16px;
}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<style>
.download-btn button {
    background: linear-gradient(135deg, #00C9A7, #92FE9D) !important;
    color: #064E3B !important;
    border-radius: 14px !important;
    font-weight: 700 !important;
    border: none !important;
}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<style>
.form-label {
    font-size: 16px;
    font-weight: 700;
    color: #4F46E5;
    margin-bottom: 6px;
}

/* Fix emoji blur */
.form-label span {
    filter: none !important;
    opacity: 1 !important;
}
</style>
""", unsafe_allow_html=True)

#marked this as remark
#st.markdown("""
#<div style="padding: 40px 0 30px 0; text-align:center;">
    #<div class="hero-title">Global Job Aggregator</div>
    #<div class="hero-subtitle">Search smarter. Apply faster.</div>
#</div>
#""", unsafe_allow_html=True)

def clear_results():
    st.session_state["jobs_df"] = None
    st.session_state["fallback_message"] = None
    st.session_state["search_triggered"] = False
# ---------- INPUT AREA ----------

# SKILLS (MANDATORY)
st.markdown('<div class="form-label">🛠 Skills *</div>', unsafe_allow_html=True)
st.caption("Update skills (comma separated). Eg: Software, Python, Java, Testing, WFM")

skills_input = st.text_input(
    "Skills",
    "",
    key="skills_input",
    label_visibility="collapsed",
    on_change=clear_results
)




skills = [s.strip() for s in skills_input.split(",") if s.strip()]


# LEVELS (OPTIONAL)
st.markdown('<div class="form-label">🎯 Levels (Optional)</div>', unsafe_allow_html=True)
st.caption("Mention levels (comma separated). Eg: Associate, Senior Developer, Manager")
levels_input = st.text_input(
    "Levels",
    "",
    key="levels_input",
    label_visibility="collapsed",
    on_change=clear_results
)





levels = [l.strip() for l in levels_input.split(",") if l.strip()]


# LOCATION
st.markdown('<div class="form-label">📍 Location (City or Remote)</div>', unsafe_allow_html=True)
st.caption("Multiple cities supported. If cities are from different countries, select those countries below. Type 'Remote' to search remote jobs.")
location_input = st.text_input(
    "Location",
    "",
    key="location_input",
    label_visibility="collapsed",
    on_change=clear_results
)






is_remote = location_input.strip().lower() == "remote"
st.markdown('<div class="form-label">🌍 Select Country</div>', unsafe_allow_html=True)

countries = st.multiselect(
    "Country",
    options=[
        "India", "United States", "United Kingdom", "United Arab Emirates",
        "Canada", "Australia", "Germany", "France", "Netherlands", "Ireland", "Singapore", "Brazil",
        "South Africa","Mexico","Poland","Belgium", "Austria","Switzerland",
        "Spain", "Italy", "Philippines"
    ],
    default=["India"],
    disabled=is_remote,
    label_visibility="collapsed",
    on_change=clear_results
)

if is_remote:
    st.info("Country selection disabled because 'Remote' location is selected.")

if not is_remote and not countries:
    st.error("Country is mandatory unless location is Remote.")
    st.stop()

posted_days = st.slider(
    "Posted within last X days",
    1,
    60,
    7,
    on_change=clear_results
)
# If the user changes inputs, clear old results so they don't see "ghost" data
# 🎁 BONUS CAREER KIT DOWNLOAD


# ---------- ACTION BAR ----------
col_run, col_toggle, col_dl = st.columns([2, 3, 2])
with col_run:
    run_btn = st.button("🚀 Run Job Search")
with col_toggle:
    sub1, sub2 = st.columns(2)

    with sub1:
        view_mode = st.toggle("Classic view", value=False)

    with sub2:
        deep_search = st.toggle(
            "🔎 Deep Search",
            value=False,
            on_change=rerun_search
        )
        st.caption("May take more time")
with col_dl:
    dl_placeholder = st.empty()

# ---------- BACKEND LOGIC ----------
def call_backend_search(payload):
    resp = requests.post(f"{BACKEND_URL}/search", json=payload, timeout=60)
    resp.raise_for_status()
    return resp.json()

@st.cache_data(ttl=1800, show_spinner=False)
def cached_backend_search(payload):
    return call_backend_search(payload)
    
# ⭐ PREVENT MULTIPLE API CALLS
if run_btn:
    if not skills:
        st.error("⚠️ At least one skill is required to run the search.")
    else:
        st.session_state["search_triggered"] = True


# CALL API WHEN RUN CLICKED OR DEEP SEARCH TOGGLED
if run_btn or st.session_state.get("rerun_needed"):
    with st.spinner("Fetching jobs..."):

        payload = {
            "skills": skills,
            "levels": levels,
            "locations": [] if is_remote else [location_input],
            "countries": countries,
            "posted_days": posted_days,
            "is_remote": is_remote,
            "page": 1,
            "page_size": 100,
            "deep_search": deep_search
        }
        
        try:
            result = cached_backend_search(payload)
            rows = result.get("rows", [])
            
            df = pd.DataFrame(rows)

            fallback = result.get("fallback", False)
            fallback_message = result.get("fallback_message")
            
            st.session_state["jobs_df"] = df
            st.session_state["fallback_message"] = fallback_message
            st.session_state["rerun_needed"] = False
        except Exception as e:
            st.error(f"Backend Error: {e}")
            df = pd.DataFrame()
            st.session_state["rerun_needed"] = False
# DISPLAY RESULTS IF THEY EXIST
df = st.session_state.get("jobs_df")

# ALWAYS load from session state

if df is None:
    st.info("Enter details and click Run Job Search.")
elif df.empty:
    st.warning("No jobs found.")
else:
    fallback_message = st.session_state.get("fallback_message")

    if "url" in df.columns:
        df = df.rename(columns={"url": "Apply"})
    # Standardize Work Mode column
    for col in ["work_mode", "WorkMode", "workMode"]:
        if col in df.columns:
            df = df.rename(columns={col: "Work Mode"})
    
    # REMOVE skill column ONLY from UI
    if "skill" in df.columns:
        df = df.drop(columns=["skill"], errors="ignore")
    if "Skill" in df.columns:
        df = df.drop(columns=["Skill"], errors="ignore")
    
    # Sort by date if available
    if "posted_date" in df.columns:
        df["posted_date"] = pd.to_datetime(df["posted_date"], errors="coerce")
        df = df.sort_values(by="posted_date", ascending=False)
    
    # Fix index numbering for display
    df = df.reset_index(drop=True)
    df.insert(0, "Sr No", range(1, len(df) + 1))
    
    st.success(f"✅ Found {len(df)} jobs")
    
    if view_mode:
        # ---------- CLASSIC TABLE VIEW ----------
        display_cols = [
            col for col in df.columns
            if col not in ["_excel", "_date"]
        ]
        
        st.dataframe(
            df[display_cols],
            use_container_width=True,
            hide_index=True,
            column_config={
                "Apply": st.column_config.LinkColumn(
                    "Apply Now",
                    display_text="Apply Now"
                )
            }
        )
    else:
        # ---------- CARD VIEW (FIXED & CLEANED) ----------
        cols = st.columns(2)
        for i, row in df.iterrows():
            # Logic to handle missing values safely
            title = row.get('title') or row.get('Title') or "Job Title"
            company = row.get('company') or row.get('Company') or "Company"
            loc = row.get('location') or row.get('Location') or "Not Specified"
            link = row.get('Apply') or row.get('apply_link') or "#"
            mode = str(row.get("Work Mode") or "On-site")
            
            # Determine Badge Color
            if "remote" in mode.lower():
                b_class = "badge-remote"
            elif "hybrid" in mode.lower():
                b_class = "badge-hybrid"
            else:
                b_class = "badge-onsite"
    
            with cols[i % 2]:
                card_html = f"""
                <div class="job-card">
                    <div class="job-title">{title}</div>
                    <div class="job-company">{company}</div>
                    <div class="job-location">📍 {loc}</div>
                    <div style="margin-top: 8px;">
                        <span class="badge {b_class}">{mode}</span>
                    </div>
                    <div class="job-actions" style="display: flex; justify-content: flex-end; margin-top: 10px;">
                        <a class="apply-btn" href="{link}" target="_blank">Apply →</a>
                    </div>
                </div>
                """
                st.markdown(card_html, unsafe_allow_html=True)
    
    # DOWNLOAD BUTTON
    with col_dl:
        st.markdown('<div class="download-btn">', unsafe_allow_html=True)
        excel_data = export_to_excel(df)
    
        dl_placeholder.download_button(
            "⬇️ Download Excel",
            excel_data,
            "job_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
        
        st.markdown('</div>', unsafe_allow_html=True)
    

# ================= GLOBAL FOOTER =================

st.markdown(
    """
    <hr style="margin-top:50px;margin-bottom:20px;">
    <div style="text-align:center;font-size:14px;color:#6b7280;">
        <b>Legal:</b>
        <a href="https://avantara.co.in/terms" target="_blank">Terms</a> |
        <a href="https://avantara.co.in/privacy" target="_blank">Privacy</a> |
        <a href="https://avantara.co.in/refund" target="_blank">Refund</a>
        <br><br>
        © 2026 Avantara
    </div>
    """,
    unsafe_allow_html=True
    
)
st.sidebar.write("VERSION TEST 123")
